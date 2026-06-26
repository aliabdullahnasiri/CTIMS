import hashlib
import os
import subprocess
import uuid
from copy import copy
from datetime import datetime as dt
from tempfile import NamedTemporaryFile, TemporaryFile

from docxtpl import DocxTemplate
from flask import current_app, send_file
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from app.blueprints.api import bp
from app.extensions.db import db
from app.models.daily_section import DailySection
from app.models.form import Form, FormType
from app.models.student import Student


def insert_row_with_style(ws, row_idx):
    ws.insert_rows(row_idx)

    source_row = row_idx - 1

    # Copy cell styles
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(row_idx, col)

        if source.has_style:
            target._style = copy(source._style)

        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.number_format = source.number_format

    # Copy row height
    ws.row_dimensions[row_idx].height = ws.row_dimensions[source_row].height

    # Copy merged cells from source row
    merged_ranges = list(ws.merged_cells.ranges)

    for merged in merged_ranges:
        min_col = merged.min_col
        max_col = merged.max_col
        min_row = merged.min_row
        max_row = merged.max_row

        if min_row == source_row and max_row == source_row:
            ws.merge_cells(
                start_row=row_idx,
                start_column=min_col,
                end_row=row_idx,
                end_column=max_col,
            )


@bp.get("/export/daily-section/<string:uid>")
def export_daily_section(uid: str):
    d = DailySection.query.filter_by(uid=uid).first_or_404()

    template_path = "app/templates/xlsx/daily_section.xlsx"

    wb = load_workbook(template_path)
    ws = wb.active

    if ws is not None:
        cols = [
            (idx, val)
            for idx in range(column_index_from_string("AE"), 0, -1)
            if (val := ws.cell(row=10, column=idx).value)
            and setattr(ws.cell(10, column=idx), "value", "") is None
        ]

        count = 1
        for idx, student in enumerate(d.students.all(), start=10):
            dct = student.to_dict()

            insert_row_with_style(ws, row := idx + 1)

            ws[f"{get_column_letter(cols[0][0])}{row}"] = count

            for col_idx, value in cols[1:]:
                ws[f"{get_column_letter(col_idx)}{row}"] = dct.get(value)

            count += 1

        ws.delete_rows(10)

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        wb.save(tmp.name)

        return send_file(
            tmp.name,
            download_name=f"daily-section-{d.uid}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


@bp.get("/export/student/<string:uid>/registration-form")
def export_student_registration_form(uid: str):
    student = Student.query.filter_by(uid=uid).first_or_404()

    if not (
        (
            form := student.forms.filter(
                Form._type == FormType.STUDENT_REGISTRATION_FORM
            )
            .order_by(getattr(Form, "id").desc())
            .first()
        )
        and form._hash
        == hashlib.sha256(
            student.updated_at.strftime("%Y-%m-%d %H:%M:%S.%f").encode("UTF-8")
        ).hexdigest()
    ):
        path = os.path.join(
            current_app.config["UPLOAD_FOLDER"],
            dt.now().strftime("%Y-%m-%d"),
        )

        os.makedirs(path, exist_ok=True)

        form: Form = Form()

        form.name = "Student Registration Form"
        form._type = FormType.STUDENT_REGISTRATION_FORM
        form._hash = hashlib.sha256(
            student.updated_at.strftime("%Y-%m-%d %H:%M:%S.%f").encode("UTF-8")
        ).hexdigest()

        form.students.add(student)
        db.session.flush()

        student.forms.add(form)

        doc = DocxTemplate("app/templates/docx/student-registration-form.docx")

        doc.render(
            {
                key.replace("-", "_"): str(value) if value is not None else ""
                for key, value in (
                    student.to_dict()
                    | {
                        "form_number": getattr(form, "id"),
                        "distribution_date": getattr(form, "display_distribution_date"),
                        "academic_year": (
                            student.daily_section.academic_year
                            if student.daily_section
                            else ""
                        ),
                    }
                ).items()
            }
        )

        doc.save(_f := os.path.join(path, filename := f"{uuid.uuid4()}.docx"))

        subprocess.run(
            [
                "libreoffice",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                path,
                _f,
            ],
            check=True,
        )

        form._path = os.path.join(path, filename.replace(".docx", ".pdf")).replace(
            "app/", ""
        )

        db.session.commit()

    return send_file(
        form._path,
        download_name=f"student_registration_{student.uid}.pdf",
        mimetype="application/pdf",
    )
